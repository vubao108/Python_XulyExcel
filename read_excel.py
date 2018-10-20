# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl import  load_workbook
import os


def readfile(filename,fromrow1,torow1,fromrow2,torow2):

    wb = load_workbook(filename)

    ws = wb.worksheets[0]
    data = {}
    for i in range(fromrow1, torow1+1):
        data[i] = {}
        for j in range(1,10):
            value = ws.cell(i,j).value
            data[i][j] = value

    ws2 = wb.worksheets[1]
    for i in range(fromrow2, torow2+1):
        data[torow1+i-1] = {}
        for j in range(1,10):
            value = ws2.cell(i,j).value
            data[torow1+i-1][j] = value
    return data

if __name__ == '__main__':
    file_name = 'C:\\Users\\vuth1\\OneDrive\\Documents\\doanhthu7892018\\doanhthu7892018\\KHDN_092018.xlsx'

    data = readfile(file_name,2,50,2,1)
    mst_count_data = {}
    mst_sum_data = {}
    mkh_count_data = {}
    mkh_sum_data = {}
    dict_mst = {}
    dict_mkh = {}
    for i in range(2,51):
        tien = int(data[i][9])
        mst = data[i][7]
        ltb = data[i][8]
        mkh = data[i][3]
        ten_tt = data[i][5]
        key01 = str(mst) + u':' + ltb
        if mst:
            dict_mst[mst] = dict_mst.get(mst, {})
            dict_mst[mst][ltb] = dict_mst[mst].get(ltb, {})
            dict_mst[mst][ltb]['soluong'] = dict_mst[mst][ltb].get('soluong',0) + 1
            dict_mst[mst][ltb]['doanhthu'] = dict_mst[mst][ltb].get('doanhthu', 0) + tien
            dict_mst[mst][ltb]['ten_tt'] = ten_tt
            '''
            key01 = str(data[i][7])  + ':' + data[i][8]
            mst_count_data[key01] = mst_count_data.get(key01, 0) + 1
            mst_sum_data[key01] = mst_sum_data.get(key01, 0) + tien
            '''
        else:
            dict_mkh[mkh] = dict_mkh.get(mkh, {})
            dict_mkh[mkh][ltb] = dict_mkh[mkh].get(ltb,{})
            dict_mkh[mkh][ltb]['soluong'] = dict_mkh[mkh][ltb].get('soluong', 0) + 1
            dict_mkh[mkh][ltb]['doanhthu'] = dict_mkh[mkh][ltb].get('doanhthu', 0) + tien
            dict_mkh[mkh][ltb]['ten_tt'] = ten_tt
            '''
            key01 = str(data[i][3])  + '' + data[i][8]
            mkh_count_data[key01] = mst_count_data.get(key01, 0) + 1
            mkh_sum_data[key01] = mst_sum_data.get(key01, 0) + tien
            '''

    bieumau = 'C:\\Users\\vuth1\\OneDrive\\Documents\\doanhthu7892018\\doanhthu7892018\\bieumau02.xlsx'
    wb = load_workbook(bieumau)
    sheet = wb.worksheets[2]
    rownum = 6
    for key_mst in dict_mst:
        sheet.cell(rownum, 4).value = key_mst
        for key_ltb in dict_mst[key_mst]:
            sheet.cell(rownum,3).value = dict_mst[key_mst][key_ltb]['ten_tt']
            if key_ltb == 'didong':
                sheet.cell(rownum,5).value = dict_mst[key_mst][key_ltb]['soluong']
                sheet.cell(rownum,6).value = dict_mst[key_mst][key_ltb]['doanhthu']
            elif key_ltb == 'dien thoai co dinh':
                sheet.cell(rownum, 7).value = dict_mst[key_mst][key_ltb]['soluong']
                sheet.cell(rownum, 8).value = dict_mst[key_mst][key_ltb]['doanhthu']
            elif key_ltb == 'Mega':
                sheet.cell(rownum, 9).value = dict_mst[key_mst][key_ltb]['soluong']
                sheet.cell(rownum, 10).value = dict_mst[key_mst][key_ltb]['doanhthu']
            elif key_ltb == 'Fiber':
                sheet.cell(rownum, 11).value = dict_mst[key_mst][key_ltb]['soluong']
                sheet.cell(rownum, 12).value = dict_mst[key_mst][key_ltb]['doanhthu']
            elif key_ltb == 'Megawan ADSL' or key_ltb == 'Megawan quang FE':
                sl_mgadsl = 0; dt_mgadsl = 0;  sl_mgq = 0; dt_mgq = 0
                if 'Megawan ADSL' in dict_mst[key_mst]:
                    sl_mgadsl = dict_mst[key_mst]['Megawan ADSL']['soluong']
                    dt_mgadsl = dict_mst[key_mst]['Megawan ADSL']['doanhthu']
                if 'Megawan quang FE' in dict_mst[key_mst]:
                    sl_mgq = dict_mst[key_mst]['Megawan quang FE']['soluong']
                    dt_mgq = dict_mst[key_mst]['Megawan quang FE']['doanhthu']

                sheet.cell(rownum, 15).value = sl_mgadsl + sl_mgq
                sheet.cell(rownum, 16).value = dt_mgadsl + dt_mgq
            elif key_ltb == 'MetroNet FE':
                sheet.cell(rownum, 17).value = dict_mst[key_mst][key_ltb]['soluong']
                sheet.cell(rownum, 18).value = dict_mst[key_mst][key_ltb]['doanhthu']
            elif key_ltb == 'MyTV':
                sheet.cell(rownum, 21).value = dict_mst[key_mst][key_ltb]['soluong']
                sheet.cell(rownum, 22).value = dict_mst[key_mst][key_ltb]['doanhthu']
        rownum = rownum + 1

    for key_mkh in dict_mkh:
        sheet.cell(rownum, 4).value = key_mkh
        for key_ltb in dict_mkh[key_mkh]:
            sheet.cell(rownum, 3).value = dict_mkh[key_mkh][key_ltb]['ten_tt']
            if key_ltb == 'didong':
                sheet.cell(rownum, 5).value = dict_mkh[key_mkh][key_ltb]['soluong']
                sheet.cell(rownum, 6).value = dict_mkh[key_mkh][key_ltb]['doanhthu']
            elif key_ltb == 'dien thoai co dinh':
                sheet.cell(rownum, 7).value = dict_mkh[key_mkh][key_ltb]['soluong']
                sheet.cell(rownum, 8).value = dict_mkh[key_mkh][key_ltb]['doanhthu']
            elif key_ltb == 'Mega':
                sheet.cell(rownum, 9).value = dict_mkh[key_mkh][key_ltb]['soluong']
                sheet.cell(rownum, 10).value = dict_mkh[key_mkh][key_ltb]['doanhthu']
            elif key_ltb == 'Fiber':
                sheet.cell(rownum, 11).value = dict_mkh[key_mkh][key_ltb]['soluong']
                sheet.cell(rownum, 12).value = dict_mkh[key_mkh][key_ltb]['doanhthu']
            elif key_ltb == 'Megawan ADSL' or key_ltb == 'Megawan quang FE':
                sl_mgadsl = 0;
                dt_mgadsl = 0;
                sl_mgq = 0;
                dt_mgq = 0
                if 'Megawan ADSL' in dict_mkh[key_mkh]:
                    sl_mgadsl = dict_mkh[key_mkh]['Megawan ADSL']['soluong']
                    dt_mgadsl = dict_mkh[key_mkh]['Megawan ADSL']['doanhthu']
                if 'Megawan quang FE' in dict_mkh[key_mkh]:
                    sl_mgq = dict_mkh[key_mkh]['Megawan quang FE']['soluong']
                    dt_mgq = dict_mkh[key_mkh]['Megawan quang FE']['doanhthu']

                sheet.cell(rownum, 15).value = sl_mgadsl + sl_mgq
                sheet.cell(rownum, 16).value = dt_mgadsl + dt_mgq
            elif key_ltb == 'MetroNet FE':
                sheet.cell(rownum, 17).value = dict_mkh[key_mkh][key_ltb]['soluong']
                sheet.cell(rownum, 18).value = dict_mkh[key_mkh][key_ltb]['doanhthu']
            elif key_ltb == 'MyTV':
                sheet.cell(rownum, 21).value = dict_mkh[key_mkh][key_ltb]['soluong']
                sheet.cell(rownum, 22).value = dict_mkh[key_mkh][key_ltb]['doanhthu']
        rownum = rownum + 1

    wb.save(bieumau)
    print dict_mst
    print dict_mkh

    #print data



'''
    empty_row = 0
    for i in range(2,200):
        print str(i-1)
        q = data[i][1]

        if q and q.strip():
            empty_row = 0
            try:
                correct_index = int(data[i][2])
            except:
                correct_index = -1
                //raw_input("ko co cau tra loi")

            a1 = data[i][3]
            a2 = data[i][4]
            a3= data[i][5]
            a4 = data[i][6]

            load_to_sqlite.insert_to_db(conn,question=q, a1=a1,a2=a2,a3=a3,a4=a4,correct_answer=correct_index, tag_name=tag_name)




        else:
            empty_row = empty_row + 1
            print tag_name + ' :' + str(i-1)
            if empty_row == 3:
                break

if __name__ == '__main__':
    all_dir = r'C:\Users\vuth1\OneDrive\Documents\On tap Thi Nang luc Ha Tinh 2018 (Moi)\On tap Thi Nang luc Ha Tinh 2018 (Moi)\Vien thong\all\remain'

    conn = load_to_sqlite.open_connection()
    for filename in os.listdir(all_dir):
        tag_name = filename.split(r'.')[0].strip()
        readfile(all_dir + r'\\' + filename, conn, tag_name=tag_name)

    conn.commit()
    conn.close()
'''
